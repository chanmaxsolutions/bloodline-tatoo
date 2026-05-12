#!/usr/bin/env python3
"""
Convert Outscraper Google Maps reviews XLSX into per-region JSON for the homepage carousel.

Usage:
  python3 scripts/generate-homepage-reviews-json.py /path/to/Outscraper-....xlsx

Writes:
  src/data/reviews/bangkok.json
  src/data/reviews/bali.json
  src/data/reviews/phuket.json
"""

from __future__ import annotations

import json
import sys
from datetime import datetime, timezone
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError as e:
    raise SystemExit("Install openpyxl: pip3 install openpyxl") from e

NAME_TO_REGION: dict[str, str] = {
    "Bloodline Tattoo Bangkok": "bangkok",
    "Bloodline Tattoo Bali - Kuta": "bali",
    "Bloodline Tattoo Patong Phuket": "phuket",
}

ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "src" / "data" / "reviews"


def is_url(value: str) -> bool:
    return value.startswith("https://") or value.startswith("http://")


def row_to_record(header: list[str], row: tuple) -> dict:
    d = {header[i]: row[i] for i in range(min(len(header), len(row)))}
    listing_name = d.get("name")
    if listing_name is None:
        raise ValueError("Row missing name")
    region = NAME_TO_REGION.get(str(listing_name).strip())
    if not region:
        raise ValueError(f"Unknown listing name: {listing_name!r}")

    rid = d.get("review_id")
    rid_str = str(rid).strip() if rid is not None else ""
    stable_id = f"{region}-{rid_str}" if rid_str else f"{region}-missing-id"

    author = str(d.get("author_title") or "").strip()
    if not author:
        raise ValueError("Row missing author_title")

    rating_raw = d.get("review_rating")
    rating = int(rating_raw) if rating_raw is not None else 0
    if rating < 1:
        rating = 1
    if rating > 5:
        rating = 5

    text = d.get("review_text")
    text_str = str(text).strip() if text is not None else ""

    ts = d.get("review_timestamp")
    date_str = ""
    if isinstance(ts, (int, float)) and ts > 0:
        date_str = datetime.fromtimestamp(float(ts), tz=timezone.utc).isoformat()
    else:
        dt = d.get("review_datetime_utc")
        date_str = str(dt).strip() if dt is not None else ""
    if not date_str:
        date_str = "1970-01-01T00:00:00.000Z"

    rec: dict = {
        "id": stable_id,
        "author": author,
        "rating": rating,
        "text": text_str,
        "date": date_str,
        "region": region,
        "source": "google",
    }

    photo = d.get("author_image")
    if photo and is_url(str(photo).strip()):
        rec["profilePhoto"] = str(photo).strip()

    link = d.get("review_link")
    if link and is_url(str(link).strip()):
        rec["reviewUrl"] = str(link).strip()

    return rec


def main() -> None:
    if len(sys.argv) < 2:
        raise SystemExit("Usage: python3 scripts/generate-homepage-reviews-json.py <path-to.xlsx>")

    xlsx = Path(sys.argv[1]).expanduser().resolve()
    if not xlsx.is_file():
        raise SystemExit(f"File not found: {xlsx}")

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter)
    header = [str(c).strip() if c is not None else "" for c in header_row]

    pools: dict[str, list[dict]] = {"bangkok": [], "bali": [], "phuket": []}
    for row in rows_iter:
        if row is None or not any(cell is not None and str(cell).strip() for cell in row):
            continue
        rec = row_to_record(header, tuple(row))
        pools[str(rec["region"])].append(rec)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    for region, items in pools.items():
        out_path = OUT_DIR / f"{region}.json"
        payload = {"reviews": items}
        out_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        print(f"Wrote {len(items)} reviews -> {out_path.relative_to(ROOT)}")

    wb.close()


if __name__ == "__main__":
    main()
